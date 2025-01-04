from io import BytesIO
from bson import ObjectId
from fastapi import HTTPException, Depends
from grpc import Status
from jose import jwt, JWTError
from auth import SECRET_KEY, ALGORITHM
from fastapi.security import OAuth2PasswordBearer
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.units import inch
from database import users_collection, admins_collection

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="login")

# Helper function to get the current user
def get_current_user(token: str = Depends(oauth2_scheme)):
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: str = payload.get("sub")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Invalid credentials")
        return user_id
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    

# Helper function to serialize MongoDB documents
def serialize_transaction(transaction):
    transaction["_id"] = str(transaction["_id"])  # Convert ObjectId to string
    transaction["user_id"] = str(transaction["user_id"])  # Convert user_id ObjectId to string if needed
    return transaction


def create_financial_report_excel(report_data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Financial Report"
    
    # Set headers style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Add report period
    ws['A1'] = "Report Period"
    ws['B1'] = f"{report_data['summary']['period_start']} to {report_data['summary']['period_end']}"
    ws.merge_cells('B1:D1')
    
    # Add Summary Section
    ws['A3'] = "Summary"
    ws.merge_cells('A3:D3')
    ws['A3'].font = Font(bold=True, size=12)
    
    summary_headers = ['Metric', 'Amount']
    for col, header in enumerate(summary_headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
    
    summary_data = [
        ('Total Income', report_data['summary']['total_income']),
        ('Total Expenses', report_data['summary']['total_expense']),
        ('Net Income', report_data['summary']['net_income']),
        ('Savings Rate', f"{report_data['summary']['savings_rate']:.2f}%")
    ]
    
    for row, (metric, value) in enumerate(summary_data, 5):
        ws.cell(row=row, column=1, value=metric)
        ws.cell(row=row, column=2, value=value)
    
    # Add Income Breakdown
    ws['A9'] = "Income Breakdown"
    ws.merge_cells('A9:D9')
    ws['A9'].font = Font(bold=True, size=12)
    
    income_headers = ['Category', 'Amount']
    for col, header in enumerate(income_headers, 1):
        cell = ws.cell(row=10, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
    
    start_row = 11
    for row, income in enumerate(report_data['income_by_category'], start_row):
        ws.cell(row=row, column=1, value=income['category'])
        ws.cell(row=row, column=2, value=income['amount'])
    
    # Add Expense Breakdown
    ws[f'A{start_row + len(report_data["income_by_category"]) + 2}'] = "Expense Breakdown"
    expense_start_row = start_row + len(report_data['income_by_category']) + 3
    ws.merge_cells(f'A{expense_start_row-1}:D{expense_start_row-1}')
    ws[f'A{expense_start_row-1}'].font = Font(bold=True, size=12)
    
    for col, header in enumerate(income_headers, 1):
        cell = ws.cell(row=expense_start_row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
    
    for row, expense in enumerate(report_data['expense_by_category'], expense_start_row + 1):
        ws.cell(row=row, column=1, value=expense['category'])
        ws.cell(row=row, column=2, value=expense['amount'])
    
    # Add Goals Summary
    goals_start_row = expense_start_row + len(report_data['expense_by_category']) + 3
    ws[f'A{goals_start_row-1}'] = "Goals Summary"
    ws.merge_cells(f'A{goals_start_row-1}:D{goals_start_row-1}')
    ws[f'A{goals_start_row-1}'].font = Font(bold=True, size=12)
    
    goals_headers = ['Goal', 'Target Amount', 'Current Amount', 'Progress']
    for col, header in enumerate(goals_headers, 1):
        cell = ws.cell(row=goals_start_row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
    
    for row, goal in enumerate(report_data['goals_summary'], goals_start_row + 1):
        ws.cell(row=row, column=1, value=goal['name'])
        ws.cell(row=row, column=2, value=goal['target_amount'])
        ws.cell(row=row, column=3, value=goal['current_amount'])
        ws.cell(row=row, column=4, value=f"{goal['progress']:.2f}%")
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
    
    return wb



def create_financial_report_pdf(report_data):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    story = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30
    )
    story.append(Paragraph("Financial Report", title_style))
    
    # Period
    period_text = f"Report Period: {report_data['summary']['period_start']} to {report_data['summary']['period_end']}"
    story.append(Paragraph(period_text, styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Summary Section
    story.append(Paragraph("Summary", styles['Heading2']))
    summary_data = [
        ['Metric', 'Amount'],
        ['Total Income', f"K{report_data['summary']['total_income']:,.2f}"],
        ['Total Expenses', f"K{report_data['summary']['total_expense']:,.2f}"],
        ['Net Income', f"K{report_data['summary']['net_income']:,.2f}"],
        ['Savings Rate', f"{report_data['summary']['savings_rate']:.2f}%"]
    ]
    
    summary_table = Table(summary_data, colWidths=[2*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 20))
    
    # Income Breakdown
    story.append(Paragraph("Income Breakdown", styles['Heading2']))
    income_data = [['Category', 'Amount']]
    for income in report_data['income_by_category']:
        income_data.append([
            income['category'],
            f"K{income['amount']:,.2f}"
        ])
    
    income_table = Table(income_data, colWidths=[2*inch, 2*inch])
    income_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.green),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.lightgreen),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(income_table)
    story.append(Spacer(1, 20))
    
    # Expense Breakdown
    story.append(Paragraph("Expense Breakdown", styles['Heading2']))
    expense_data = [['Category', 'Amount']]
    for expense in report_data['expense_by_category']:
        expense_data.append([
            expense['category'],
            f"K{expense['amount']:,.2f}"
        ])
    
    expense_table = Table(expense_data, colWidths=[2*inch, 2*inch])
    expense_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.red),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.pink),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(expense_table)
    story.append(Spacer(1, 20))
    
    # Goals Summary
    if report_data['goals_summary']:
        story.append(Paragraph("Goals Summary", styles['Heading2']))
        goals_data = [['Goal', 'Target', 'Current', 'Progress']]
        for goal in report_data['goals_summary']:
            goals_data.append([
                goal['name'],
                f"K{goal['target_amount']:,.2f}",
                f"K{goal['current_amount']:,.2f}",
                f"{goal['progress']:.2f}%"
            ])
        
        goals_table = Table(goals_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
        goals_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.blue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightblue),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(goals_table)
    
    # Build PDF
    doc.build(story)
    buffer.seek(0)
    return buffer


async def get_current_admin(current_user: str = Depends(get_current_user)):
    user = admins_collection.find_one({"_id": ObjectId(current_user)})
    if not user or user.get("role") != "admin":
        raise HTTPException(
            status_code=Status.HTTP_403_FORBIDDEN,
            detail="Admin access required"
        )
    return current_user
